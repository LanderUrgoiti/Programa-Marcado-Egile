import threading
import socket
import tkinter as tk
import tkinter.messagebox as messagebox
from PIL import ImageTk, Image
import time
import re
import select
import pyodbc as po
import pyautogui
import pygetwindow
import pyperclip
import datetime
import os
import win32api
import openpyxl
import shutil
import qrcode
from openpyxl.drawing.image import Image


nombre_archivo='Hoja registro ' + 'PRUEBA' +'.xlsx'
nombre_impresora='Ineo + 3350 Marcadora NGVS'
# Busca el archivo en la carpeta actual
carpeta=r"\\srv5\Maquinas\Documentacion NGV's\HOJAS DE REGISTRO UNIFICADAS"
ruta_archivo = os.path.join(r"\\srv5\Maquinas\Documentacion NGV's\HOJAS DE REGISTRO UNIFICADAS", nombre_archivo)

for filename in os.listdir(carpeta):
    if str('PRUEBA') in filename:
        ruta_archivo = os.path.join(carpeta, filename)
        ruta_programa = f'TNC:\manual\{of}.H'
        #ruta_hexadecimal = ruta_programa.encode("utf-8").hex()
        # Abre el archivo Excel
        libro = openpyxl.load_workbook(ruta_archivo)
        hoja = libro.active

        # Generar el código QR
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=3,
            border=1,
        )
        qr.add_data(ruta_programa)
        qr.make(fit=True)

        # Crear la imagen del código QR
        img = qr.make_image(fill='black', back_color='white')
        img_path = "qrcode.png"
        img.save(img_path)
        img_to_insert = Image(img_path)
        hoja.add_image(img_to_insert, 'J5')
        
        hoja.cell(row=1, column=10, value='333333')
        #hoja.cell(row=1, column=10, value=ruta_programa)
        libro.save(r"\\srv5\Maquinas\Documentacion NGV's\HOJAS DE REGISTRO UNIFICADAS" + 'Temp.xlsx')

        win32api.ShellExecute(
            0,
            "print",
            r"\\srv5\Maquinas\Documentacion NGV's\HOJAS DE REGISTRO UNIFICADAS" + "Temp.xlsx",
            f'/d:"{nombre_impresora}"',
            ".",
            0
        )
        libro.close()
        time.sleep(3)

